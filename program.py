import openpyxl

# Nombre del archivo de origen y archivo de destino
archivo_origen = "datos_origen.xlsx"
archivo_destino = "datos_destino.xlsx"

# Cargar el archivo de origen
libro_origen = openpyxl.load_workbook(archivo_origen)

# Obtener la hoja de cálculo activa
hoja_origen = libro_origen.active

# Crear un nuevo libro de Excel
libro_destino = openpyxl.Workbook()
hoja_destino = libro_destino.active

# Recorrer las filas del archivo de origen y copiar los datos al archivo de destino
for fila_origen in hoja_origen.iter_rows(values_only=True):
    # Crear una nueva fila para el archivo de destino
    fila_destino = []
    for valor in fila_origen:
        # Eliminar los espacios al inicio de cada valor y agregarlos a la nueva fila
        if valor is not None and valor != "":
            fila_destino.append(str(valor).strip())
            print("La celda contiene dato:", valor)
        else:
            print("La celda esta vacia")
            continue
    
    # Copiar la fila completa al archivo de destino
    hoja_destino.append(fila_destino)

# Guardar el archivo de destino
libro_destino.save(archivo_destino)

# Cerrar los libros
libro_origen.close()
libro_destino.close()

print("Datos copiados correctamente al archivo de destino.")